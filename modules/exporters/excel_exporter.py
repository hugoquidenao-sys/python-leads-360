"""
exporters/excel_exporter.py - Exportación de leads a Excel

Genera leads.xlsx con formato profesional, colores por clasificación
y todas las columnas requeridas.
"""
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR
from utils.logger import logger


# ── Colores por clasificación ─────────────────────────────────────────────────────
COLOR_MAP = {
    "HOT":  "FF4444",   # Rojo
    "WARM": "FF9900",   # Naranja
    "COLD": "4472C4",   # Azul
}
HEADER_COLOR = "1F3864"   # Azul oscuro corporativo


def export_leads_excel(leads: list[dict], filename: str = "leads.xlsx") -> Path:
    """
    Exporta la lista de leads a un archivo Excel formateado.

    Args:
        leads:    lista de dicts del formato get_full_leads()
        filename: nombre del archivo de salida

    Returns:
        Path al archivo generado
    """
    if not leads:
        logger.warning("No hay leads para exportar")
        return Path()

    rows = _build_rows(leads)
    output_path = OUTPUT_DIR / filename

    # Crear DataFrame
    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False, engine="openpyxl")

    # Aplicar formato
    _format_workbook(output_path, rows)

    logger.info(f"✓ Excel exportado: {output_path} ({len(rows)} leads)")
    return output_path


# ── Construcción de filas ─────────────────────────────────────────────────────────

def _build_rows(leads: list[dict]) -> list[dict]:
    rows = []
    for lead in leads:
        c    = lead["company"]
        a    = lead.get("analysis", {})
        pain = lead.get("pain_points", [])
        recs = lead.get("recommendations", [])

        pain_text = " | ".join(p["pain"] for p in pain[:4])
        services  = " | ".join(r["service"] for r in recs[:3])
        emails    = ", ".join(a.get("emails", [])[:2])
        phones_web = ", ".join(a.get("phones", [])[:2])
        phone     = phones_web or c.get("phone", "")

        rows.append({
            "Empresa":               c.get("name", ""),
            "Sitio Web":             c.get("website", ""),
            "Correo":                emails,
            "Teléfono":              phone,
            "Industria":             c.get("category", ""),
            "Ciudad":                c.get("city", ""),
            "Score":                 lead.get("score", 0),
            "Clasificación":         lead.get("classification", "COLD"),
            "Pain Points":           pain_text,
            "Servicios Recomendados": services,
            "N° Señales":            len(lead.get("signals", [])),
            "Dirección":             c.get("address", ""),
            "Fecha análisis":        datetime.now().strftime("%d/%m/%Y"),
        })

    # Ordenar HOT → WARM → COLD, luego por score desc
    order = {"HOT": 0, "WARM": 1, "COLD": 2}
    rows.sort(key=lambda r: (order.get(r["Clasificación"], 3), -r["Score"]))
    return rows


# ── Formato Excel ─────────────────────────────────────────────────────────────────

def _format_workbook(path: Path, rows: list[dict]) -> None:
    """Aplica formato profesional al workbook."""
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Leads"

    n_cols = ws.max_column
    n_rows = ws.max_row

    # ── Header row ─────────────────────────────────────────────────────────────
    header_font    = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill    = PatternFill("solid", fgColor=HEADER_COLOR)
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border    = Border(
        bottom=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="DDDDDD"),
    )

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = thin_border

    ws.row_dimensions[1].height = 30

    # ── Data rows ──────────────────────────────────────────────────────────────
    score_col  = _find_col(ws, "Score")
    class_col  = _find_col(ws, "Clasificación")
    center_align = Alignment(vertical="center")
    wrap_align   = Alignment(vertical="top", wrap_text=True)

    for row_num in range(2, n_rows + 1):
        classification = ""
        if class_col:
            classification = ws.cell(row=row_num, column=class_col).value or ""

        row_fill = PatternFill(
            "solid",
            fgColor=_row_color(classification, row_num)
        )

        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            # Columnas de texto largo
            if col in {_find_col(ws, "Pain Points"), _find_col(ws, "Servicios Recomendados")}:
                cell.alignment = wrap_align
            else:
                cell.alignment = center_align

        # Color suave de fondo por clasificación
        for col in range(1, n_cols + 1):
            ws.cell(row=row_num, column=col).fill = PatternFill(
                "solid", fgColor=_row_color(classification, row_num)
            )

        ws.row_dimensions[row_num].height = 40

    # ── Anchos de columna ──────────────────────────────────────────────────────
    col_widths = {
        "Empresa": 30, "Sitio Web": 28, "Correo": 28,
        "Teléfono": 16, "Industria": 22, "Ciudad": 14,
        "Score": 8, "Clasificación": 14, "Pain Points": 45,
        "Servicios Recomendados": 40, "N° Señales": 10,
        "Dirección": 30, "Fecha análisis": 14,
    }

    for col in range(1, n_cols + 1):
        header_val = ws.cell(row=1, column=col).value
        width = col_widths.get(str(header_val), 18)
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def _find_col(ws, header: str) -> int | None:
    """Encuentra el número de columna por nombre de encabezado."""
    for cell in ws[1]:
        if cell.value == header:
            return cell.column
    return None


def _row_color(classification: str, row_num: int) -> str:
    """Retorna color de fondo suave según clasificación y fila alternada."""
    base_colors = {
        "HOT":  ("FFE8E8", "FFCCCC"),
        "WARM": ("FFF4E6", "FFE8CC"),
        "COLD": ("E8F0FF", "CCE0FF"),
    }
    alt = row_num % 2
    colors = base_colors.get(classification, ("FFFFFF", "F5F5F5"))
    return colors[alt]
